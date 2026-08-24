"""Leitura e escrita de planilhas (.xlsx/.csv) para importar/exportar em
massa a distribuição de lucro por sócio. Puro Python — sem dependência de UI
— pra dar pra testar a lógica de parsing isoladamente."""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COLUNAS_DISTRIBUICAO = ["CPF", "Sócio", "Valor Distribuído", "Pró-labore", "IRRF"]

COLUNAS_CADASTRO = [
    "Nº Chamada", "Empresa", "CNPJ", "Capital Social", "Qtd. Cotas da Empresa",
    "Sócio", "CPF/CNPJ do Sócio", "Tipo (física/jurídica)",
    "% Capital do Sócio", "Cotas do Sócio", "Data de Entrada", "Data de Saída",
    "Ano Base", "Valor Distribuído", "Pró-labore", "IRRF",
]


def exportar_modelo_distribuicao(caminho: Path, linhas: list[dict]) -> None:
    """Gera um .xlsx pronto pra preencher: uma linha por sócio informado, já
    com CPF e nome preenchidos — só falta digitar o valor distribuído (e,
    se houver, pró-labore/IRRF)."""
    workbook = openpyxl.Workbook()
    aba = workbook.active
    aba.title = "Distribuição"
    aba.append(COLUNAS_DISTRIBUICAO)
    for linha in linhas:
        aba.append(
            [
                linha.get("cpf", ""),
                linha.get("nome", ""),
                linha.get("valor_distribuido") or 0,
                linha.get("pro_labore") or 0,
                linha.get("irrf") or 0,
            ]
        )
    for coluna, largura in zip("ABCDE", (20, 32, 18, 14, 14)):
        aba.column_dimensions[coluna].width = largura
    workbook.save(caminho)


def importar_distribuicao(caminho: Path) -> list[dict]:
    """Lê uma planilha (.xlsx ou .csv) com colunas CPF / Sócio / Valor
    Distribuído / Pró-labore / IRRF (nessa ordem ou não, identificadas pelo
    cabeçalho — as duas últimas são opcionais) e retorna uma lista de
    {"cpf", "nome", "valor_distribuido", "pro_labore", "irrf"}. Não valida
    contra o banco — isso é responsabilidade de quem chama."""
    caminho = Path(caminho)
    linhas_brutas = _ler_csv(caminho) if caminho.suffix.lower() == ".csv" else _ler_xlsx(caminho)
    linhas_brutas = [linha for linha in linhas_brutas if any(c not in (None, "") for c in linha)]

    if not linhas_brutas:
        return []

    cabecalho = [str(c or "").strip().lower() for c in linhas_brutas[0]]
    indice_cpf = _indice_coluna(cabecalho, ["cpf"])
    indice_nome = _indice_coluna(cabecalho, ["sócio", "socio", "nome"])
    indice_valor = _indice_coluna(cabecalho, ["valor distribuído", "valor distribuido", "valor"])
    indice_pro_labore = _indice_coluna(cabecalho, ["pró-labore", "pro-labore", "pro labore", "prolabore"])
    indice_irrf = _indice_coluna(cabecalho, ["irrf"])

    if indice_valor is None:
        raise ValueError(
            'Não encontrei a coluna "Valor Distribuído" na planilha. '
            "Use o modelo exportado pelo sistema (botão \"Exportar modelo\") pra garantir o formato certo."
        )
    if indice_cpf is None and indice_nome is None:
        raise ValueError('A planilha precisa ter uma coluna "CPF" ou "Sócio" pra identificar cada linha.')

    def numero_opcional(linha: list, indice: int | None, numero_linha: int, rotulo: str) -> float:
        if indice is None or indice >= len(linha):
            return 0.0
        try:
            return _para_numero(linha[indice])
        except ValueError:
            raise ValueError(f'Linha {numero_linha}: {rotulo} inválido "{linha[indice]}".') from None

    resultado = []
    for numero_linha, linha in enumerate(linhas_brutas[1:], start=2):
        cpf = str(linha[indice_cpf]).strip() if indice_cpf is not None and indice_cpf < len(linha) and linha[indice_cpf] else ""
        nome = str(linha[indice_nome]).strip() if indice_nome is not None and indice_nome < len(linha) and linha[indice_nome] else ""
        valor_bruto = linha[indice_valor] if indice_valor < len(linha) else None
        try:
            valor = _para_numero(valor_bruto)
        except ValueError:
            raise ValueError(f'Linha {numero_linha}: valor inválido "{valor_bruto}".') from None
        if not cpf and not nome:
            continue
        resultado.append(
            {
                "cpf": cpf,
                "nome": nome,
                "valor_distribuido": valor,
                "pro_labore": numero_opcional(linha, indice_pro_labore, numero_linha, "pró-labore"),
                "irrf": numero_opcional(linha, indice_irrf, numero_linha, "IRRF"),
            }
        )
    return resultado


def _indice_coluna(cabecalho: list[str], nomes_possiveis: list[str]) -> int | None:
    for nome in nomes_possiveis:
        if nome in cabecalho:
            return cabecalho.index(nome)
    return None


def _para_numero(valor) -> float:
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace("R$", "").strip()
    if "," in texto and texto.rfind(",") > texto.rfind("."):
        texto = texto.replace(".", "").replace(",", ".")
    return float(texto)


def _ler_csv(caminho: Path) -> list[list[str]]:
    with open(caminho, newline="", encoding="utf-8-sig") as arquivo:
        amostra = arquivo.read(2048)
        arquivo.seek(0)
        delimitador = ";" if amostra.count(";") >= amostra.count(",") else ","
        return list(csv.reader(arquivo, delimiter=delimitador))


def _ler_xlsx(caminho: Path) -> list[list]:
    workbook = openpyxl.load_workbook(caminho, data_only=True)
    aba = workbook.active
    return [list(linha) for linha in aba.iter_rows(values_only=True)]


def exportar_modelo_cadastro(caminho: Path, linhas: list[dict] | None = None) -> None:
    """Gera um .xlsx pronto pra preencher com o cadastro em massa de empresas,
    sócios e vínculos: uma linha por (empresa, sócio). Sem linhas, sai só com
    o cabeçalho — útil pra cadastrar do zero. Com linhas (cadastro atual),
    serve de referência/edição. Data de Saída, Ano Base, Valor Distribuído,
    Pró-labore e IRRF são opcionais — só preenche quem tiver saído da
    sociedade ou tiver uma distribuição daquele ano pra lançar junto."""
    workbook = openpyxl.Workbook()
    aba = workbook.active
    aba.title = "Cadastro"
    aba.append(COLUNAS_CADASTRO)
    for linha in linhas or []:
        aba.append(
            [
                linha.get("numero_chamada", ""),
                linha.get("empresa_nome", ""),
                linha.get("cnpj", ""),
                linha.get("capital_social") or 0,
                linha.get("quantidade_cotas") or 0,
                linha.get("socio_nome", ""),
                linha.get("socio_cpf", ""),
                linha.get("tipo_pessoa", "fisica"),
                linha.get("percentual_capital") or 0,
                linha.get("cotas_socio") or 0,
                linha.get("data_entrada", ""),
                linha.get("data_saida", ""),
                linha.get("ano_base", ""),
                linha.get("valor_distribuido") or 0,
                linha.get("pro_labore") or 0,
                linha.get("irrf") or 0,
            ]
        )
    for coluna, largura in zip(
        "ABCDEFGHIJKLMNOP", (10, 28, 20, 16, 16, 28, 20, 14, 14, 14, 14, 14, 10, 16, 14, 14)
    ):
        aba.column_dimensions[coluna].width = largura
    workbook.save(caminho)


def importar_cadastro(caminho: Path) -> list[dict]:
    """Lê uma planilha (.xlsx ou .csv) de cadastro em massa — colunas
    identificadas pelo cabeçalho, na ordem de COLUNAS_CADASTRO ou não — e
    retorna uma lista de dicts com os campos crus (ainda não validados nem
    casados contra o banco; isso é responsabilidade de quem chama)."""
    caminho = Path(caminho)
    linhas_brutas = _ler_csv(caminho) if caminho.suffix.lower() == ".csv" else _ler_xlsx(caminho)
    linhas_brutas = [linha for linha in linhas_brutas if any(c not in (None, "") for c in linha)]

    if not linhas_brutas:
        return []

    cabecalho = [str(c or "").strip().lower() for c in linhas_brutas[0]]
    idx = {
        "numero_chamada": _indice_coluna(cabecalho, ["nº chamada", "numero chamada", "n° chamada", "chamada"]),
        "empresa_nome": _indice_coluna(cabecalho, ["empresa", "nome da empresa", "razão social"]),
        "cnpj": _indice_coluna(cabecalho, ["cnpj"]),
        "capital_social": _indice_coluna(cabecalho, ["capital social", "capital"]),
        "quantidade_cotas": _indice_coluna(cabecalho, ["qtd. cotas da empresa", "cotas da empresa", "quantidade de cotas"]),
        "socio_nome": _indice_coluna(cabecalho, ["sócio", "socio", "nome do sócio", "nome do socio"]),
        "socio_cpf": _indice_coluna(cabecalho, ["cpf/cnpj do sócio", "cpf/cnpj do socio", "cpf do sócio", "cpf do socio", "cpf"]),
        "tipo_pessoa": _indice_coluna(cabecalho, ["tipo (física/jurídica)", "tipo (fisica/juridica)", "tipo"]),
        "percentual_capital": _indice_coluna(cabecalho, ["% capital do sócio", "% capital do socio", "percentual capital", "% capital"]),
        "cotas_socio": _indice_coluna(cabecalho, ["cotas do sócio", "cotas do socio"]),
        "data_entrada": _indice_coluna(cabecalho, ["data de entrada", "data entrada"]),
        "data_saida": _indice_coluna(cabecalho, ["data de saída", "data de saida", "data saída", "data saida"]),
        "ano_base": _indice_coluna(cabecalho, ["ano base", "ano_base", "ano"]),
        "valor_distribuido": _indice_coluna(cabecalho, ["valor distribuído", "valor distribuido", "valor"]),
        "pro_labore": _indice_coluna(cabecalho, ["pró-labore", "pro-labore", "pro labore", "prolabore"]),
        "irrf": _indice_coluna(cabecalho, ["irrf"]),
    }

    if idx["empresa_nome"] is None:
        raise ValueError(
            'Não encontrei a coluna "Empresa" na planilha. '
            "Use o modelo exportado pelo sistema (botão \"Exportar modelo\") pra garantir o formato certo."
        )
    if idx["socio_nome"] is None:
        raise ValueError('Não encontrei a coluna "Sócio" na planilha.')

    def texto(linha: list, chave: str) -> str:
        i = idx[chave]
        if i is None or i >= len(linha) or linha[i] is None:
            return ""
        return str(linha[i]).strip()

    def numero_opcional(linha: list, chave: str, numero_linha: int, rotulo: str) -> float:
        i = idx[chave]
        if i is None or i >= len(linha) or linha[i] in (None, ""):
            return 0.0
        try:
            return _para_numero(linha[i])
        except ValueError:
            raise ValueError(f'Linha {numero_linha}: {rotulo} inválido "{linha[i]}".') from None

    resultado = []
    for numero_linha, linha in enumerate(linhas_brutas[1:], start=2):
        empresa_nome = texto(linha, "empresa_nome")
        socio_nome = texto(linha, "socio_nome")
        if not empresa_nome or not socio_nome:
            continue

        try:
            capital_social = numero_opcional(linha, "capital_social", numero_linha, "capital social")
            quantidade_cotas = numero_opcional(linha, "quantidade_cotas", numero_linha, "quantidade de cotas")
            percentual_capital = numero_opcional(linha, "percentual_capital", numero_linha, "percentual de capital")
            cotas_socio = numero_opcional(linha, "cotas_socio", numero_linha, "cotas do sócio")
            valor_distribuido = numero_opcional(linha, "valor_distribuido", numero_linha, "valor distribuído")
            pro_labore = numero_opcional(linha, "pro_labore", numero_linha, "pró-labore")
            irrf = numero_opcional(linha, "irrf", numero_linha, "IRRF")
        except ValueError:
            raise

        i_data = idx["data_entrada"]
        data_bruta = linha[i_data] if i_data is not None and i_data < len(linha) else None
        try:
            data_entrada = _para_data(data_bruta)
        except ValueError:
            raise ValueError(f'Linha {numero_linha}: data de entrada inválida "{data_bruta}".') from None

        i_saida = idx["data_saida"]
        saida_bruta = linha[i_saida] if i_saida is not None and i_saida < len(linha) else None
        try:
            data_saida = _para_data_opcional(saida_bruta)
        except ValueError:
            raise ValueError(f'Linha {numero_linha}: data de saída inválida "{saida_bruta}".') from None

        i_ano = idx["ano_base"]
        ano_bruto = linha[i_ano] if i_ano is not None and i_ano < len(linha) else None
        ano_base = None
        if ano_bruto not in (None, ""):
            try:
                ano_base = int(_para_numero(ano_bruto))
            except ValueError:
                raise ValueError(f'Linha {numero_linha}: ano base inválido "{ano_bruto}".') from None

        tipo_bruto = texto(linha, "tipo_pessoa").lower()
        tipo_pessoa = "juridica" if tipo_bruto.startswith(("j", "pj")) else "fisica"

        resultado.append(
            {
                "numero_chamada": texto(linha, "numero_chamada"),
                "empresa_nome": empresa_nome,
                "cnpj": texto(linha, "cnpj"),
                "capital_social": capital_social,
                "quantidade_cotas": quantidade_cotas,
                "socio_nome": socio_nome,
                "socio_cpf": texto(linha, "socio_cpf"),
                "tipo_pessoa": tipo_pessoa,
                "percentual_capital": percentual_capital,
                "data_saida": data_saida,
                "ano_base": ano_base,
                "valor_distribuido": valor_distribuido,
                "pro_labore": pro_labore,
                "irrf": irrf,
                "cotas_socio": cotas_socio,
                "data_entrada": data_entrada,
            }
        )
    return resultado


def _converter_data(valor) -> str:
    """Aceita datetime/date (openpyxl já converte células de data) ou texto
    em dd/mm/aaaa ou aaaa-mm-dd. Não aceita vazio — quem chama decide o que
    fazer nesse caso (ver _para_data e _para_data_opcional)."""
    import datetime as _dt

    if isinstance(valor, _dt.datetime):
        return valor.date().isoformat()
    if isinstance(valor, _dt.date):
        return valor.isoformat()
    texto = str(valor).strip()
    if "/" in texto:
        dia, mes, ano = texto.split("/")
        return _dt.date(int(ano), int(mes), int(dia)).isoformat()
    return _dt.date.fromisoformat(texto).isoformat()


def _para_data(valor) -> str:
    """Vazio vira hoje — toda linha de vínculo precisa de uma data de
    entrada, então na ausência de uma informada assume-se a data atual."""
    import datetime as _dt

    if valor is None or valor == "":
        return _dt.date.today().isoformat()
    return _converter_data(valor)


def _para_data_opcional(valor) -> str | None:
    """Vazio vira None — ausência de data de saída significa que o sócio
    continua ativo, bem diferente de "saiu hoje"."""
    if valor is None or valor == "":
        return None
    return _converter_data(valor)


def exportar_relatorio_excel(
    caminho: Path, titulo: str, cabecalho: list[str], linhas: list[list], aba_nome: str = "Relatório"
) -> None:
    """Exporta um relatório tabular genérico (título + cabeçalho + linhas)
    com formatação básica — cabeçalho destacado, colunas com largura
    ajustada ao conteúdo, primeira linha congelada pra rolar com contexto."""
    workbook = openpyxl.Workbook()
    aba = workbook.active
    aba.title = aba_nome[:31]

    aba.append([titulo])
    if cabecalho:
        aba.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cabecalho))
    aba["A1"].font = Font(size=14, bold=True, color="1B2A41")
    aba.append([])

    linha_cabecalho = 3
    aba.append(cabecalho)
    for celula in aba[linha_cabecalho]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="1B2A41")

    for linha in linhas:
        aba.append(linha)

    aba.freeze_panes = f"A{linha_cabecalho + 1}"
    for indice_coluna in range(1, len(cabecalho) + 1):
        letra = get_column_letter(indice_coluna)
        maior = max(
            (len(str(aba.cell(row=r, column=indice_coluna).value or "")) for r in range(linha_cabecalho, aba.max_row + 1)),
            default=10,
        )
        aba.column_dimensions[letra].width = min(max(maior + 2, 10), 42)

    workbook.save(caminho)
